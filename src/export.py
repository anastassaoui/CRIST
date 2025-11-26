"""
Export functionality for Excel and PDF reports.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
import io
import config


def create_excel_report(evaporator_results, crystallizer_results, optimization_results, integration_results):
    """
    Create formatted Excel workbook with all simulation results.

    Returns:
    --------
    BytesIO : Excel file in memory
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14, color="1F4E78")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "EVAPORATION AND CRYSTALLIZATION PROCESS"
    ws_summary['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws_summary['A2'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    row = 4

    # Evaporator summary
    if evaporator_results and 'summary' in evaporator_results:
        ws_summary[f'A{row}'] = "EVAPORATOR RESULTS"
        ws_summary[f'A{row}'].font = title_font
        row += 1

        summary = evaporator_results['summary']
        data = [
            ['Number of Effects', summary['n_effects'], '-'],
            ['Final Concentration', f"{summary['final_concentration']*100:.2f}", '%'],
            ['Steam Consumption', f"{summary['steam_consumption']:.0f}", 'kg/h'],
            ['Steam Economy', f"{summary['steam_economy']:.2f}", 'kg vapor/kg steam'],
            ['Total Heat Transfer Area', f"{summary['total_area']:.1f}", 'm²'],
        ]

        for item in data:
            ws_summary[f'A{row}'] = item[0]
            ws_summary[f'B{row}'] = item[1]
            ws_summary[f'C{row}'] = item[2]
            row += 1
        row += 1

    # Crystallizer summary
    if crystallizer_results:
        ws_summary[f'A{row}'] = "CRYSTALLIZER RESULTS"
        ws_summary[f'A{row}'].font = title_font
        row += 1

        data = [
            ['Cooling Strategy', crystallizer_results.get('strategy', 'N/A'), '-'],
            ['Mean Crystal Size (L50)', f"{crystallizer_results.get('L50_microns', 0):.1f}", 'μm'],
            ['Coefficient of Variation', f"{crystallizer_results.get('CV', 0)*100:.1f}", '%'],
            ['Mass Yield', f"{crystallizer_results.get('yield', 0)*100:.1f}", '%'],
            ['Final Concentration', f"{crystallizer_results.get('final_concentration', 0):.1f}", 'g/100g'],
        ]

        for item in data:
            ws_summary[f'A{row}'] = item[0]
            ws_summary[f'B{row}'] = item[1]
            ws_summary[f'C{row}'] = item[2]
            row += 1
        row += 1

    # Optimization summary
    if optimization_results and optimization_results.get('status') == 'optimal':
        ws_summary[f'A{row}'] = "OPTIMIZATION RESULTS"
        ws_summary[f'A{row}'].font = title_font
        row += 1

        data = [
            ['Status', optimization_results['status'], '-'],
            ['Steam Consumption', f"{optimization_results['steam_consumption']:.0f}", 'kg/h'],
            ['Total Area', f"{optimization_results.get('total_area', 0):.1f}", 'm²'],
            ['Steam Economy', f"{optimization_results.get('steam_economy', 0):.2f}", '-'],
            ['Annualized Cost', f"{optimization_results['objective_value']:,.0f}", '€/year'],
        ]

        for item in data:
            ws_summary[f'A{row}'] = item[0]
            ws_summary[f'B{row}'] = item[1]
            ws_summary[f'C{row}'] = item[2]
            row += 1

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15

    # Sheet 2: Evaporator Details
    if evaporator_results and 'effects' in evaporator_results:
        ws_evap = wb.create_sheet("Evaporator Details")

        df = pd.DataFrame(evaporator_results['effects'])

        # Write header
        ws_evap['A1'] = "EVAPORATOR EFFECT-BY-EFFECT RESULTS"
        ws_evap['A1'].font = title_font
        ws_evap.merge_cells('A1:J1')

        # Write dataframe
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_evap.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 3:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                cell.border = border

        # Adjust column widths
        for col_idx in range(1, ws_evap.max_column + 1):
            ws_evap.column_dimensions[get_column_letter(col_idx)].width = 15

    # Sheet 3: Crystallizer Details
    if crystallizer_results:
        ws_cryst = wb.create_sheet("Crystallizer Details")

        ws_cryst['A1'] = "CRYSTALLIZATION BATCH RESULTS"
        ws_cryst['A1'].font = title_font

        row = 3
        data = [
            ['Parameter', 'Value', 'Unit'],
            ['Initial Concentration', f"{crystallizer_results.get('initial_concentration', 0):.1f}", 'g/100g'],
            ['Final Concentration', f"{crystallizer_results.get('final_concentration', 0):.1f}", 'g/100g'],
            ['Initial Temperature', f"{crystallizer_results.get('initial_temperature', 0):.1f}", '°C'],
            ['Final Temperature', f"{crystallizer_results.get('final_temperature', 0):.1f}", '°C'],
            ['Batch Duration', f"{crystallizer_results.get('duration', 0)/3600:.1f}", 'hours'],
            ['Mean Crystal Size (L50)', f"{crystallizer_results.get('L50_microns', 0):.1f}", 'μm'],
            ['CV', f"{crystallizer_results.get('CV', 0)*100:.1f}", '%'],
            ['Mass Yield', f"{crystallizer_results.get('yield', 0)*100:.1f}", '%'],
        ]

        for item in data:
            for c_idx, value in enumerate(item, start=1):
                cell = ws_cryst.cell(row=row, column=c_idx, value=value)

                if row == 3:  # Header
                    cell.fill = header_fill
                    cell.font = header_font

                cell.border = border
            row += 1

        # Add equipment sizing if available
        if 'sizing' in crystallizer_results:
            sizing = crystallizer_results['sizing']
            row += 1
            ws_cryst.cell(row=row, column=1, value="EQUIPMENT SIZING").font = Font(bold=True, size=12, color="1F4E78")
            row += 2

            sizing_data = [
                ['Parameter', 'Value', 'Unit'],
                ['Tank Diameter', f"{sizing.get('tank_diameter', 0):.2f}", 'm'],
                ['Total Volume (with freeboard)', f"{sizing.get('total_volume', 0):.2f}", 'm³'],
                ['Impeller Diameter', f"{sizing['agitation'].get('impeller_diameter', 0):.2f}", 'm'],
                ['Motor Power', f"{sizing['agitation'].get('power', 0):.2f}", 'kW'],
                ['Rotation Speed', f"{sizing['agitation'].get('rotation_speed', 0):.0f}", 'rpm'],
                ['Tip Speed', f"{sizing['agitation'].get('tip_speed', 0):.2f}", 'm/s'],
                ['Reynolds Number', f"{sizing['agitation'].get('reynolds_number', 0):.0f}", '-'],
                ['Heat Exchange Area', f"{sizing['cooling'].get('area', 0):.2f}", 'm²'],
                ['Cooling Power', f"{sizing['cooling'].get('Q_avg', 0):.2f}", 'kW'],
                ['Total Heat Removed', f"{sizing['cooling'].get('Q_total', 0):.0f}", 'kJ'],
                ['Batches per Day', f"{sizing['residence'].get('batches_per_day', 0):.2f}", '-'],
                ['Annual Production', f"{sizing['residence'].get('annual_throughput', 0):.0f}", 'tonnes/year'],
            ]

            for item in sizing_data:
                for c_idx, value in enumerate(item, start=1):
                    cell = ws_cryst.cell(row=row, column=c_idx, value=value)

                    if item == sizing_data[0]:  # Header
                        cell.fill = header_fill
                        cell.font = header_font

                    cell.border = border
                row += 1

        ws_cryst.column_dimensions['A'].width = 35
        ws_cryst.column_dimensions['B'].width = 20
        ws_cryst.column_dimensions['C'].width = 15

    # Sheet 4: Heat Integration
    if integration_results:
        ws_integ = wb.create_sheet("Heat Integration")

        ws_integ['A1'] = "HEAT INTEGRATION ANALYSIS"
        ws_integ['A1'].font = title_font

        row = 3
        data = [
            ['Parameter', 'Value', 'Unit'],
            ['Heat Available (Hot Streams)', f"{integration_results.get('Q_available', 0)/1000:.0f}", 'kW'],
            ['Heat Required (Cold Streams)', f"{integration_results.get('Q_required', 0)/1000:.0f}", 'kW'],
            ['Recoverable Heat', f"{integration_results.get('Q_recoverable', 0)/1000:.0f}", 'kW'],
            ['Heat Recovery Percentage', f"{integration_results.get('heat_recovery_percent', 0):.1f}", '%'],
            ['Heat Exchanger Area', f"{integration_results.get('A_heat_exchanger', 0):.1f}", 'm²'],
        ]

        for item in data:
            for c_idx, value in enumerate(item, start=1):
                cell = ws_integ.cell(row=row, column=c_idx, value=value)

                if row == 3:
                    cell.fill = header_fill
                    cell.font = header_font

                cell.border = border
            row += 1

        ws_integ.column_dimensions['A'].width = 35
        ws_integ.column_dimensions['B'].width = 20
        ws_integ.column_dimensions['C'].width = 15

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_pdf_report(evaporator_results, crystallizer_results, optimization_results, integration_results):
    """
    Create formatted PDF report with all simulation results.

    Returns:
    --------
    BytesIO : PDF file in memory
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                          rightMargin=0.75*inch, leftMargin=0.75*inch,
                          topMargin=0.75*inch, bottomMargin=0.75*inch)

    # Container for the 'Flowable' objects
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=12,
        spaceBefore=12
    )

    # Title
    title = Paragraph("EVAPORATION AND CRYSTALLIZATION PROCESS", title_style)
    elements.append(title)

    subtitle = Paragraph(f"<b>Simulation Report</b><br/>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        styles['Normal'])
    elements.append(subtitle)
    elements.append(Spacer(1, 0.3*inch))

    # Project Information
    elements.append(Paragraph("Project Information", heading_style))

    project_data = [
        ['Parameter', 'Value'],
        ['Project', 'Sugar Production - Evaporation & Crystallization'],
        ['Feed Flow Rate', f"{config.FEED_FLOW_RATE} kg/h"],
        ['Feed Concentration', f"{config.FEED_CONCENTRATION*100:.1f}%"],
        ['Target Concentration', f"{config.TARGET_CONCENTRATION*100:.1f}%"],
    ]

    project_table = Table(project_data, colWidths=[3*inch, 3*inch])
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(project_table)
    elements.append(Spacer(1, 0.3*inch))

    # Evaporator Results
    if evaporator_results and 'summary' in evaporator_results:
        elements.append(Paragraph("1. EVAPORATOR RESULTS", heading_style))

        summary = evaporator_results['summary']

        evap_data = [
            ['Parameter', 'Value', 'Unit'],
            ['Number of Effects', str(summary['n_effects']), '-'],
            ['Final Concentration', f"{summary['final_concentration']*100:.2f}", '%'],
            ['Steam Consumption', f"{summary['steam_consumption']:.0f}", 'kg/h'],
            ['Steam Economy', f"{summary['steam_economy']:.2f}", 'kg vapor/kg steam'],
            ['Total Heat Transfer Area', f"{summary['total_area']:.1f}", 'm²'],
        ]

        evap_table = Table(evap_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        evap_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(evap_table)
        elements.append(Spacer(1, 0.2*inch))

        # Effect-by-effect details
        if 'effects' in evaporator_results:
            elements.append(Paragraph("Effect-by-Effect Details", styles['Heading3']))

            df = pd.DataFrame(evaporator_results['effects'])

            # Select key columns
            effect_data = [['Effect', 'L out (kg/h)', 'V out (kg/h)', 'Conc. (%)', 'Temp (°C)', 'Area (m²)']]

            for _, row in df.iterrows():
                effect_data.append([
                    str(row['effect']),
                    f"{row['L_out']:.0f}",
                    f"{row['V_out']:.0f}",
                    f"{row['x_out']*100:.1f}",
                    f"{row['T_boiling']:.1f}",
                    f"{row['A']:.1f}"
                ])

            effect_table = Table(effect_data, colWidths=[1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            effect_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(effect_table)

        elements.append(Spacer(1, 0.3*inch))

    # Crystallizer Results
    if crystallizer_results:
        elements.append(Paragraph("2. CRYSTALLIZER RESULTS", heading_style))

        cryst_data = [
            ['Parameter', 'Value', 'Unit'],
            ['Cooling Strategy', crystallizer_results.get('strategy', 'N/A').capitalize(), '-'],
            ['Mean Crystal Size (L50)', f"{crystallizer_results.get('L50_microns', 0):.1f}", 'μm'],
            ['Coefficient of Variation', f"{crystallizer_results.get('CV', 0)*100:.1f}", '%'],
            ['Mass Yield', f"{crystallizer_results.get('yield', 0)*100:.1f}", '%'],
            ['Final Concentration', f"{crystallizer_results.get('final_concentration', 0):.1f}", 'g/100g'],
        ]

        cryst_table = Table(cryst_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        cryst_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(cryst_table)
        elements.append(Spacer(1, 0.2*inch))

        # Equipment sizing if available
        if 'sizing' in crystallizer_results:
            elements.append(Paragraph("Equipment Sizing", styles['Heading3']))

            sizing = crystallizer_results['sizing']

            sizing_data = [
                ['Parameter', 'Value', 'Unit'],
                ['Tank Diameter', f"{sizing.get('tank_diameter', 0):.2f}", 'm'],
                ['Total Volume', f"{sizing.get('total_volume', 0):.2f}", 'm³'],
                ['Impeller Diameter', f"{sizing['agitation'].get('impeller_diameter', 0):.2f}", 'm'],
                ['Motor Power', f"{sizing['agitation'].get('power', 0):.2f}", 'kW'],
                ['Rotation Speed', f"{sizing['agitation'].get('rotation_speed', 0):.0f}", 'rpm'],
                ['Heat Exchange Area', f"{sizing['cooling'].get('area', 0):.2f}", 'm²'],
                ['Cooling Power', f"{sizing['cooling'].get('Q_avg', 0):.2f}", 'kW'],
                ['Batches per Day', f"{sizing['residence'].get('batches_per_day', 0):.2f}", '-'],
                ['Annual Production', f"{sizing['residence'].get('annual_throughput', 0):.0f}", 'tonnes/year'],
            ]

            sizing_table = Table(sizing_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
            sizing_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(sizing_table)

        elements.append(Spacer(1, 0.3*inch))

    # Page break
    elements.append(PageBreak())

    # Optimization Results
    if optimization_results and optimization_results.get('status') == 'optimal':
        elements.append(Paragraph("3. OPTIMIZATION RESULTS", heading_style))

        opt_data = [
            ['Parameter', 'Value', 'Unit'],
            ['Optimization Status', optimization_results['status'].capitalize(), '-'],
            ['Steam Consumption', f"{optimization_results['steam_consumption']:.0f}", 'kg/h'],
            ['Total Heat Transfer Area', f"{optimization_results.get('total_area', 0):.1f}", 'm²'],
            ['Steam Economy', f"{optimization_results.get('steam_economy', 0):.2f}", '-'],
            ['Annualized Cost', f"{optimization_results['objective_value']:,.0f}", '€/year'],
        ]

        opt_table = Table(opt_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        opt_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightyellow),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(opt_table)
        elements.append(Spacer(1, 0.3*inch))

    # Heat Integration Results
    if integration_results:
        elements.append(Paragraph("4. HEAT INTEGRATION ANALYSIS", heading_style))

        integ_data = [
            ['Parameter', 'Value', 'Unit'],
            ['Heat Available (Hot Streams)', f"{integration_results.get('Q_available', 0)/1000:.0f}", 'kW'],
            ['Heat Required (Cold Streams)', f"{integration_results.get('Q_required', 0)/1000:.0f}", 'kW'],
            ['Recoverable Heat', f"{integration_results.get('Q_recoverable', 0)/1000:.0f}", 'kW'],
            ['Heat Recovery Percentage', f"{integration_results.get('heat_recovery_percent', 0):.1f}", '%'],
            ['Heat Exchanger Area', f"{integration_results.get('A_heat_exchanger', 0):.1f}", 'm²'],
        ]

        integ_table = Table(integ_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        integ_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFE6CC')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(integ_table)

    # Conclusions
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("5. CONCLUSIONS", heading_style))

    conclusions = []

    if evaporator_results and 'summary' in evaporator_results:
        economy = evaporator_results['summary']['steam_economy']
        if economy > 3.0:
            conclusions.append("• Evaporator performance is excellent with high steam economy.")
        elif economy > 2.0:
            conclusions.append("• Evaporator performance is good and meets targets.")
        else:
            conclusions.append("• Evaporator steam economy could be improved through optimization.")

    if crystallizer_results:
        cv = crystallizer_results.get('CV', 0)
        if cv < 0.25:
            conclusions.append("• Crystal size distribution is uniform with low CV.")
        elif cv < 0.35:
            conclusions.append("• Crystal size distribution is acceptable.")
        else:
            conclusions.append("• Crystal size distribution shows high variation - consider optimization.")

    if integration_results:
        recovery = integration_results.get('heat_recovery_percent', 0)
        if recovery > 50:
            conclusions.append(f"• Significant energy savings ({recovery:.1f}%) possible through heat integration.")
        elif recovery > 30:
            conclusions.append(f"• Moderate energy savings ({recovery:.1f}%) achievable through heat recovery.")

    if not conclusions:
        conclusions.append("• Complete all simulations for comprehensive analysis.")

    for conclusion in conclusions:
        elements.append(Paragraph(conclusion, styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    return buffer
